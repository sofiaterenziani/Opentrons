"""Compute per-well k_cat from a BioTek 384-well DCPIP-linked ADH kinetic export.

Pipeline
--------
    1. Parse the plate-reader .xlsx into per-well absorbance-vs-time traces.
    2. Convert Abs -> [DCPIP] using the standard-curve slope `eps_app` (AU/µM)
       from ``dcpip_standard_curve.py``.
    3. Fit an initial-velocity slope v0 to each well (linear region only).
    4. k_cat = v0 (µM/s) / [E] (µM) -> units of s^-1.
    5. Render an editable-SVG 16x24 grid of per-well traces (with k_cat printed
       in each panel) + a k_cat heat map, and dump a stats table.
    6. Optional --average of triplicate plates (well-wise mean/std of k_cat).
    7. Optional --compare across (protein, pH) conditions -> summary CSV +
       side-by-side heat maps.

Plate layout (test plate; ``PlateLayout`` for real plates)
----------------------------------------------------------
    Test plate:
        * All 24 columns  -> Ca in all wells
        * Rows A..O       -> ethanol substrate
        * Row P cols 1-12 -> water (negative control)
        * Row P cols 13-24 -> TCEP (positive control; reduces DCPIP directly)
    Screen plate (``PlateLayout`` / ``make_screen_plate_layout``):
        * Columns 1..N_METALS -> metals (see ``METALS``)
        * Rows A..O           -> alcohols (see ``ALCOHOLS``)
        * Row P               -> no substrate

Usage
-----
    # Single plate:
    python kcat_plate.py test/260731_full_plate_test.xlsx \\
        --standard-curve test/260731_dcpip_standard_curve.xlsx

    # Triplicate average:
    python kcat_plate.py plate1.xlsx plate2.xlsx plate3.xlsx --average \\
        --eps 0.006035 --label ADH1_pH7

    # Cross-condition comparison (needs a manifest JSON, see --help).
    python kcat_plate.py --compare manifest.json --eps 0.006035
"""

from __future__ import annotations

import argparse
import csv
import json
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable

import matplotlib.pyplot as plt
import numpy as np
import openpyxl


# =============================================================================
# Editable configuration
# =============================================================================

DEFAULT_ENZYME_CONC_UM: float = 1.0
DEFAULT_EPS_APP_AU_PER_UM: float = 0.00603483  # from dcpip_standard_curve on 260731
DEFAULT_OUTPUT_SUBDIR: str = "outputs"

# Metals for the future plates, ordered by atomic number.
# 24 metals -> col 24 is Lu (no room for a no-metal control column); row P (no-substrate)
# is then the only within-plate negative. Sr excluded per protocol.
METALS: list[str] = [
    "Al",  # 13
    "Ca",  # 20
    "Sc",  # 21
    "Mn",  # 25
    "Fe",  # 26
    "Co",  # 27
    "Ni",  # 28
    "Cu",  # 29
    "Zn",  # 30
    "Y",   # 39
    "La",  # 57
    "Ce",  # 58
    "Pr",  # 59
    "Nd",  # 60
    "Sm",  # 62
    "Eu",  # 63
    "Gd",  # 64
    "Tb",  # 65
    "Dy",  # 66
    "Ho",  # 67
    "Er",  # 68
    "Tm",  # 69
    "Yb",  # 70
    "Lu",  # 71
]

# 15 alcohol substrates (placeholder list -- edit for real plates).
ALCOHOLS: list[str] = [
    "methanol",
    "ethanol",
    "1-propanol",
    "1-butanol",
    "3-methyl-1-butanol",
    "1-pentanol",
    "2-methyl-1-butanol",
    "1-hexanol",
    "1,5-pentanediol",
    "2-phenylethanol",
    "4-hydroxybenzyl alcohol",
    "vanillin",
    "protocatechuic acid",
    "vanillyl alcohol",
    "vanillic acid",
]

# Control markers appearing in the well-condition dict.
NO_METAL = "no_metal"
NO_SUBSTRATE = "no_substrate"
WATER_CONTROL = "water"     # negative control (buffer only in place of substrate)
TCEP_CONTROL = "tcep"       # positive control (reduces DCPIP directly)

PLATE_ROWS: list[str] = list("ABCDEFGHIJKLMNOP")   # 16
PLATE_COLS: list[int] = list(range(1, 25))          # 24

# Nature-panel style (full-column figure: double-column width, ~half page tall).
NATURE_RC = {
    "svg.fonttype": "none",
    "pdf.fonttype": 42,
    "font.family": "sans-serif",
    "font.sans-serif": ["Helvetica", "Arial", "DejaVu Sans"],
    "font.size": 6,
    "axes.labelsize": 6,
    "axes.titlesize": 6,
    "xtick.labelsize": 6,
    "ytick.labelsize": 6,
    "legend.fontsize": 6,
    "axes.linewidth": 0.5,
    "xtick.major.width": 0.5,
    "ytick.major.width": 0.5,
    "xtick.major.size": 2.0,
    "ytick.major.size": 2.0,
    "axes.spines.top": False,
    "axes.spines.right": False,
}
NATURE_PANEL_INCHES = (7.08, 5.20)  # ~180 x 132 mm (double-column full-column figure)


# =============================================================================
# Data classes
# =============================================================================

@dataclass
class PlateTraces:
    """Kinetic traces for a 384-well plate."""

    times_s: np.ndarray                              # shape (T,)
    traces: dict[str, np.ndarray]                    # well_id -> (T,) absorbance
    temperature_c: np.ndarray | None = None          # shape (T,), optional
    source: Path | None = None

    @property
    def well_ids(self) -> list[str]:
        return list(self.traces.keys())


@dataclass
class PlateLayout:
    """Which condition each well contains."""

    metal: dict[str, str]        # well -> metal symbol or NO_METAL
    substrate: dict[str, str]    # well -> substrate name or NO_SUBSTRATE
    description: str = ""

    def condition(self, well: str) -> tuple[str, str]:
        return self.metal.get(well, NO_METAL), self.substrate.get(well, NO_SUBSTRATE)

    def is_blank(self, well: str) -> bool:
        m, s = self.condition(well)
        return m == NO_METAL and s == NO_SUBSTRATE

    def is_no_substrate(self, well: str) -> bool:
        return self.substrate.get(well, NO_SUBSTRATE) == NO_SUBSTRATE

    def is_no_metal(self, well: str) -> bool:
        return self.metal.get(well, NO_METAL) == NO_METAL

    def is_water_control(self, well: str) -> bool:
        return self.substrate.get(well) == WATER_CONTROL

    def is_tcep_control(self, well: str) -> bool:
        return self.substrate.get(well) == TCEP_CONTROL


@dataclass
class WellFit:
    """Initial-velocity fit for one well."""

    well: str
    v0_abs_per_s: float       # slope of A vs t (negative for DCPIP reduction)
    intercept: float
    r_squared: float
    window_s: tuple[float, float]
    n_points: int
    kcat_s: float             # k_cat in s^-1, sign-corrected (positive = turnover)


@dataclass
class PlateResult:
    traces: PlateTraces
    layout: PlateLayout
    fits: dict[str, WellFit]
    eps_app: float
    enzyme_uM: float
    stats: dict[str, float] = field(default_factory=dict)
    excluded: set[str] = field(default_factory=set)  # auto-flagged "no enzyme" wells
    baseline_trace: np.ndarray | None = None         # water-control mean A(t) used for correction
    rotate_mode: str = "auto"                        # CLI rotate flag that was used
    rotation_detected: bool = False                  # auto-detector thought the plate was rotated 180°
    rotation_applied: bool = False                   # whether well IDs were physically remapped

    def kcat_grid(self) -> np.ndarray:
        """Return a 16x24 array of k_cat (NaN if fit failed or well excluded)."""
        grid = np.full((len(PLATE_ROWS), len(PLATE_COLS)), np.nan)
        for i, r in enumerate(PLATE_ROWS):
            for j, c in enumerate(PLATE_COLS):
                w = f"{r}{c}"
                fit = self.fits.get(w)
                if fit is not None:
                    grid[i, j] = fit.kcat_s
        return grid


# =============================================================================
# CLI / main entry point (defined up-front; body relies on the helpers below)
# =============================================================================

def _build_argparser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    p.add_argument("xlsx", nargs="*", type=Path,
                   help="One or more plate-reader .xlsx files. "
                        "One = single plate; multiple with --average = triplicate.")
    p.add_argument("--standard-curve", type=Path, default=None,
                   help="DCPIP standard-curve .xlsx (fit on the fly for eps_app).")
    p.add_argument("--eps", type=float, default=None,
                   help=f"Override eps_app (AU/µM). Default {DEFAULT_EPS_APP_AU_PER_UM}.")
    p.add_argument("--enzyme-uM", type=float, default=DEFAULT_ENZYME_CONC_UM,
                   help=f"Enzyme concentration in µM (default {DEFAULT_ENZYME_CONC_UM}).")
    p.add_argument("--window-s", type=float, default=300.0,
                   help="Linear-fit window width in seconds (default 300 s).")
    p.add_argument("--layout", choices=("test", "screen"), default="test",
                   help="Plate layout preset: 'test' = uniform Ca+EtOH; 'screen' = metals x alcohols grid.")
    p.add_argument("--label", type=str, default="",
                   help="Label used in output filenames.")
    p.add_argument("--average", action="store_true",
                   help="Treat multiple input files as replicates; output averaged heat map + std.")
    p.add_argument("--compare", type=Path, default=None,
                   help="JSON manifest {condition_name: [plate.xlsx, ...]} for cross-condition compare.")
    p.add_argument("--rotate", choices=("auto", "off", "180"), default="auto",
                   help="Handle plates loaded rotated 180 degrees. 'auto' (default) "
                        "detects via the control-row signature and remaps well IDs; "
                        "'off' disables any rotation; '180' forces rotation.")
    return p


def _default_outdir(xlsx: Path) -> Path:
    return xlsx.resolve().parent.parent / DEFAULT_OUTPUT_SUBDIR


def _run_single(xlsx: Path, args, layout: "PlateLayout", eps_app: float) -> "PlateResult":
    result = analyze_plate(xlsx, eps_app=eps_app, enzyme_uM=args.enzyme_uM,
                           layout=layout, window_s=args.window_s,
                           rotate=args.rotate)
    outdir = _default_outdir(xlsx)
    stem = args.label or xlsx.stem
    plot_traces_grid(result, outdir / f"{stem}_traces_grid.pdf", title=f"{stem} traces")
    plot_kcat_heatmap(result, outdir / f"{stem}_kcat_heatmap.pdf", title=f"{stem} k$_{{cat}}$")
    dump_stats(result, outdir / f"{stem}_stats.txt")
    print(f"[{xlsx.name}]")
    for k, v in result.stats.items():
        print(f"  {k:>22s} = {v:.4g}")
    return result


def main() -> None:
    args = _build_argparser().parse_args()
    eps_app = load_eps_app(args.eps, args.standard_curve)

    if args.compare is not None:
        manifest = json.loads(args.compare.read_text())
        layout = make_test_plate_layout() if args.layout == "test" else make_screen_plate_layout()
        cond_grids: dict[str, np.ndarray] = {}
        outdir = _default_outdir(args.compare)
        for name, files in manifest.items():
            plate_results = [
                analyze_plate(Path(f), eps_app=eps_app, enzyme_uM=args.enzyme_uM,
                              layout=layout, window_s=args.window_s,
                              rotate=args.rotate)
                for f in files
            ]
            mean_grid, _std = average_plate_results(plate_results)
            cond_grids[name] = mean_grid
        compare_conditions(cond_grids, layout,
                           outdir / "compare_conditions.pdf",
                           outdir / "compare_conditions.csv")
        print(f"Wrote comparison across {len(cond_grids)} conditions to {outdir}/compare_conditions.*")
        return

    if not args.xlsx:
        raise SystemExit("Provide at least one .xlsx (or use --compare).")

    layout = make_test_plate_layout() if args.layout == "test" else make_screen_plate_layout()

    if args.average and len(args.xlsx) > 1:
        results = [_run_single(x, args, layout, eps_app) for x in args.xlsx]
        mean_grid, std_grid = average_plate_results(results)
        outdir = _default_outdir(args.xlsx[0])
        stem = args.label or "average"
        plot_averaged_heatmaps(mean_grid, std_grid, layout,
                               outdir / f"{stem}_avg_heatmaps.pdf",
                               title=f"{stem}: mean/std of {len(results)} replicates")
        print(f"Wrote averaged heat maps to {outdir}/{stem}_avg_heatmaps.pdf")
    else:
        for xlsx in args.xlsx:
            _run_single(xlsx, args, layout, eps_app)


# =============================================================================
# Parser: BioTek 384-well kinetic export
# =============================================================================

def _time_to_seconds(t) -> float:
    """Convert BioTek 'Time' cell to seconds; supports datetime.time or numeric."""
    if hasattr(t, "hour"):
        return t.hour * 3600.0 + t.minute * 60.0 + t.second + getattr(t, "microsecond", 0) / 1e6
    if isinstance(t, (int, float)):
        # Some exports use fractional days.
        return float(t) * 86400.0 if 0.0 <= t < 2.0 else float(t)
    raise TypeError(f"Unrecognized time cell: {t!r}")


def _is_well_id(v) -> bool:
    return (
        isinstance(v, str)
        and 2 <= len(v) <= 3
        and v[0].upper() in "ABCDEFGHIJKLMNOP"
        and v[1:].isdigit()
        and 1 <= int(v[1:]) <= 24
    )


def parse_kinetic_plate(xlsx_path: str | Path, sheet: str | None = None) -> PlateTraces:
    """Load all per-well traces from a BioTek 384-well kinetic export.

    The export writes multiple stacked blocks (usually 4 x 96 wells). Each block starts
    with a header row [None, 'Time', 'T° ...', 'A1', 'A2', ...] followed by data rows
    until the next blank row.
    """
    xlsx_path = Path(xlsx_path)
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]

    all_rows = list(ws.iter_rows(values_only=True))

    times_ref: list[float] | None = None
    temperature: list[float] | None = None
    traces: dict[str, np.ndarray] = {}

    i = 0
    while i < len(all_rows):
        row = all_rows[i]
        # Look for a header row containing "Time" and at least one well id.
        try:
            time_col = row.index("Time")
        except ValueError:
            i += 1
            continue
        well_cols: list[tuple[int, str]] = [
            (idx, v.upper()) for idx, v in enumerate(row) if _is_well_id(v)
        ]
        if not well_cols:
            i += 1
            continue

        # Temperature column, if present, is between Time and the first well.
        first_well_idx = well_cols[0][0]
        temp_col: int | None = None
        for idx in range(time_col + 1, first_well_idx):
            cell = row[idx]
            if isinstance(cell, str) and "°" in cell:
                temp_col = idx
                break

        block_times: list[float] = []
        block_temp: list[float] = []
        block_data: dict[str, list[float]] = {w: [] for _, w in well_cols}

        j = i + 1
        while j < len(all_rows):
            drow = all_rows[j]
            if drow[time_col] is None:
                break
            try:
                block_times.append(_time_to_seconds(drow[time_col]))
            except TypeError:
                break
            if temp_col is not None and isinstance(drow[temp_col], (int, float)):
                block_temp.append(float(drow[temp_col]))
            for idx, well in well_cols:
                v = drow[idx]
                block_data[well].append(float(v) if isinstance(v, (int, float)) else np.nan)
            j += 1

        block_t = np.asarray(block_times)
        if times_ref is None:
            times_ref = list(block_t)
            if temp_col is not None and block_temp:
                temperature = list(block_temp)

        for w, vals in block_data.items():
            arr = np.asarray(vals, dtype=float)
            # Reference block may be slightly longer/shorter -- pad/truncate to match.
            if len(arr) != len(times_ref):
                m = min(len(arr), len(times_ref))
                arr = arr[:m]
            traces[w] = arr

        i = j + 1

    if times_ref is None:
        raise ValueError(f"No kinetic blocks found in {xlsx_path}")

    return PlateTraces(
        times_s=np.asarray(times_ref),
        traces=traces,
        temperature_c=np.asarray(temperature) if temperature else None,
        source=xlsx_path,
    )


# =============================================================================
# Initial-velocity fit
# =============================================================================

def _linear_fit(x: np.ndarray, y: np.ndarray) -> tuple[float, float, float]:
    """Return (slope, intercept, r_squared) for a simple OLS line."""
    n = len(x)
    if n < 3:
        return np.nan, np.nan, np.nan
    xm, ym = x.mean(), y.mean()
    dx, dy = x - xm, y - ym
    sxx = float(np.sum(dx * dx))
    if sxx == 0:
        return np.nan, np.nan, np.nan
    slope = float(np.sum(dx * dy) / sxx)
    intercept = float(ym - slope * xm)
    resid = y - (slope * x + intercept)
    ss_res = float(np.sum(resid * resid))
    ss_tot = float(np.sum(dy * dy))
    r2 = 1.0 - ss_res / ss_tot if ss_tot > 0 else np.nan
    return slope, intercept, r2


def fit_initial_velocity(
    t: np.ndarray,
    a: np.ndarray,
    *,
    window_s: float = 300.0,
    min_r2: float = 0.90,
    search_frac: float = 0.5,
) -> tuple[float, float, float, tuple[float, float], int]:
    """Fit v0 (Abs/s) with a window that always starts at t[0].

    Grows the window from a 4-point minimum up to ``min(window_s, search_frac * duration)``
    and picks the *longest* window with R^2 >= ``min_r2`` (so the fit stays inside
    the linear region as far as it holds). Falls back to the minimum window if
    nothing meets the R^2 gate.
    """
    finite = np.isfinite(a)
    t = t[finite]
    a = a[finite]
    if len(t) < 5:
        return np.nan, np.nan, np.nan, (np.nan, np.nan), 0

    t_end_cap = t[0] + min(window_s, search_frac * (t[-1] - t[0]))
    best: tuple[float, ...] | None = None
    fallback: tuple[float, ...] | None = None

    for i1 in range(4, len(t) + 1):
        if t[i1 - 1] > t_end_cap:
            break
        slope, intercept, r2 = _linear_fit(t[:i1], a[:i1])
        if not np.isfinite(slope):
            continue
        packed = (slope, intercept, r2, (t[0], t[i1 - 1]), i1)
        if fallback is None:
            fallback = packed
        if np.isnan(r2) or r2 >= min_r2:
            best = packed  # keep overwriting so we end on the longest passing window

    return best if best is not None else (fallback if fallback is not None
                                          else (np.nan, np.nan, np.nan, (np.nan, np.nan), 0))


def _water_baseline_trace(
    traces: PlateTraces,
    layout: PlateLayout,
) -> np.ndarray | None:
    """Mean A(t) across water (negative control) wells; falls back to fully blank
    wells if none are marked as water controls. Returns None if neither exists."""
    baseline_wells = [w for w in traces.traces if layout.is_water_control(w)]
    if not baseline_wells:
        baseline_wells = [w for w in traces.traces if layout.is_blank(w)]
    if not baseline_wells:
        return None
    stacked = np.vstack([traces.traces[w] for w in baseline_wells])
    return np.nanmean(stacked, axis=0)


def compute_well_fits(
    traces: PlateTraces,
    layout: PlateLayout,
    eps_app: float,
    enzyme_uM: float,
    *,
    window_s: float = 300.0,
) -> dict[str, WellFit]:
    """Fit each well and convert slope -> k_cat (s^-1).

    Baseline drift is estimated from the water (negative) control wells if
    present, otherwise from wells marked as fully blank. The mean baseline
    trace is subtracted from every well before the initial-velocity fit, so
    non-zero drift in the assay background does not bias v0.
    """
    fits: dict[str, WellFit] = {}
    baseline_trace = _water_baseline_trace(traces, layout)

    for well, a in traces.traces.items():
        # Controls are not baseline-corrected: water IS the baseline, and TCEP
        # provides the positive-control raw signal for Z'.
        if (baseline_trace is None
                or layout.is_water_control(well)
                or layout.is_tcep_control(well)):
            a_use = a
        else:
            a_use = a - baseline_trace
        slope, intercept, r2, window, n_pts = fit_initial_velocity(
            traces.times_s, a_use, window_s=window_s
        )
        if np.isfinite(slope):
            # DCPIP is reduced -> Abs falls -> slope < 0. Positive turnover rate:
            v0_uM_per_s = -slope / eps_app
            kcat = v0_uM_per_s / enzyme_uM if enzyme_uM > 0 else np.nan
        else:
            kcat = np.nan
        fits[well] = WellFit(
            well=well,
            v0_abs_per_s=slope,
            intercept=intercept,
            r_squared=r2,
            window_s=window,
            n_points=n_pts,
            kcat_s=kcat,
        )
    return fits


# =============================================================================
# Plate layout helpers
# =============================================================================

def make_test_plate_layout(
    metal: str = "Ca",
    substrate: str = "ethanol",
) -> PlateLayout:
    """Test plate: Ca + ethanol in all rows A..O across all 24 columns.

    Row P holds controls (metal still Ca):
        * P1..P12  -> water (negative control)
        * P13..P24 -> TCEP (positive control; reduces DCPIP directly)
    """
    m: dict[str, str] = {}
    s: dict[str, str] = {}
    for r in PLATE_ROWS:
        for c in PLATE_COLS:
            w = f"{r}{c}"
            m[w] = metal
            if r == "P":
                s[w] = WATER_CONTROL if c <= 12 else TCEP_CONTROL
            else:
                s[w] = substrate
    return PlateLayout(
        metal=m, substrate=s,
        description=(f"test plate: {metal} everywhere; rows A-O = {substrate}; "
                     f"row P cols 1-12 = water (neg), cols 13-24 = TCEP (pos)"),
    )


def make_screen_plate_layout(
    metals: Iterable[str] = METALS,
    alcohols: Iterable[str] = ALCOHOLS,
) -> PlateLayout:
    """Layout for real screening plates.

    Rows A..O -> alcohols; row P -> no substrate.
    Cols 1..len(metals) -> metals; any leftover cols -> no-metal control.
    If ``len(metals) >= 24`` there is no room for a no-metal column and col 24
    holds the 24th metal instead. Truncates silently if lists exceed row/col count.
    """
    metals = list(metals)[: len(PLATE_COLS)]
    alcohols = list(alcohols)[: len(PLATE_ROWS) - 1]  # leave row P for "no substrate"
    m: dict[str, str] = {}
    s: dict[str, str] = {}
    for i, r in enumerate(PLATE_ROWS):
        for j, c in enumerate(PLATE_COLS):
            w = f"{r}{c}"
            m[w] = metals[j] if j < len(metals) else NO_METAL
            s[w] = alcohols[i] if (r != "P" and i < len(alcohols)) else NO_SUBSTRATE
    no_metal_cols = len(PLATE_COLS) - len(metals)
    return PlateLayout(
        metal=m, substrate=s,
        description=(f"screen: {len(metals)} metals x {len(alcohols)} alcohols; "
                     f"{no_metal_cols} no-metal col(s), rowP=no substrate"),
    )


# =============================================================================
# QC statistics
# =============================================================================

def _raw_zprime_control_wells(result: PlateResult) -> tuple[list[str], list[str]]:
    """Use the fixed test-plate control wells for raw Z' calculations.

    This keeps the raw assay QC comparable across layouts even when an experimental
    screen layout reassigns row P to no-substrate wells. The "test" control pair
    is the canonical plate QC and should remain constant across runs.
    """
    test_layout = make_test_plate_layout()
    water_wells = [w for w in result.traces.traces if test_layout.is_water_control(w)]
    tcep_wells = [w for w in result.traces.traces if test_layout.is_tcep_control(w)]

    if water_wells and tcep_wells:
        return water_wells, tcep_wells

    # Fallback to the active layout if the file itself actually contains explicit
    # water/TCEP wells (e.g. a real test plate or custom layout with controls).
    return (
        [w for w in result.fits if result.layout.is_water_control(w)],
        [w for w in result.fits if result.layout.is_tcep_control(w)],
    )


def compute_plate_stats(result: PlateResult) -> dict[str, float]:
    """Simple plate-wide QC: mean/std/CV of positive, no-metal and no-substrate wells,
    plus Z' factor comparing positive vs no-substrate wells.
    """
    layout = result.layout
    kcats: dict[str, float] = {w: f.kcat_s for w, f in result.fits.items()}

    pos = np.array([v for w, v in kcats.items()
                    if not layout.is_no_substrate(w) and not layout.is_no_metal(w)
                    and not layout.is_water_control(w) and not layout.is_tcep_control(w)
                    and np.isfinite(v)])
    no_sub = np.array([v for w, v in kcats.items()
                       if layout.is_no_substrate(w) and not layout.is_no_metal(w)
                       and np.isfinite(v)])
    no_met = np.array([v for w, v in kcats.items()
                       if layout.is_no_metal(w) and not layout.is_no_substrate(w)
                       and np.isfinite(v)])

    def _safe(fn, arr):
        return float(fn(arr)) if arr.size else float("nan")

    stats = {
        "n_pos": float(pos.size),
        "pos_mean_kcat": _safe(np.mean, pos),
        "pos_std_kcat": _safe(np.std, pos),
        "pos_cv_pct": _safe(lambda x: 100 * np.std(x) / np.mean(x), pos) if pos.size and np.mean(pos) != 0 else float("nan"),
        "n_no_sub": float(no_sub.size),
        "no_sub_mean_kcat": _safe(np.mean, no_sub),
        "no_sub_std_kcat": _safe(np.std, no_sub),
        "n_no_metal": float(no_met.size),
        "no_metal_mean_kcat": _safe(np.mean, no_met),
        "no_metal_std_kcat": _safe(np.std, no_met),
    }
    # Z' factor: 1 - 3(sigma_p + sigma_n) / |mu_p - mu_n|
    if pos.size and no_sub.size:
        denom = abs(stats["pos_mean_kcat"] - stats["no_sub_mean_kcat"])
        if denom > 0:
            stats["z_prime"] = 1 - 3 * (stats["pos_std_kcat"] + stats["no_sub_std_kcat"]) / denom

    # Raw-absorbance Z' is computed with the fixed test-plate control wells so the
    # score remains comparable across screen/test layouts.
    times = result.traces.times_s
    zprime_mask = times <= Z_PRIME_WINDOW_S
    if not np.any(zprime_mask):
        zprime_mask = np.zeros_like(times, dtype=bool)
        zprime_mask[: min(3, len(times))] = True

    def _zprime_mean(w: str) -> float:
        a = result.traces.traces.get(w)
        if a is None:
            return float("nan")
        return float(np.nanmean(a[zprime_mask]))

    water_wells, tcep_wells = _raw_zprime_control_wells(result)
    tcep = np.array([_zprime_mean(w) for w in tcep_wells])
    tcep = tcep[np.isfinite(tcep)]
    water = np.array([_zprime_mean(w) for w in water_wells])
    water = water[np.isfinite(water)]
    if tcep.size and water.size:
        stats["n_tcep"] = float(tcep.size)
        stats["n_water"] = float(water.size)
        stats["zprime_window_s"] = float(Z_PRIME_WINDOW_S)
        stats["tcep_mean_zprime_abs"] = float(np.mean(tcep))
        stats["tcep_std_zprime_abs"] = float(np.std(tcep))
        stats["water_mean_zprime_abs"] = float(np.mean(water))
        stats["water_std_zprime_abs"] = float(np.std(water))
        denom_raw = abs(stats["water_mean_zprime_abs"] - stats["tcep_mean_zprime_abs"])
        if denom_raw > 0:
            stats["z_prime_raw"] = 1 - 3 * (
                stats["water_std_zprime_abs"] + stats["tcep_std_zprime_abs"]
            ) / denom_raw
    return stats


# =============================================================================
# Plotting
# =============================================================================

def plot_traces_grid(result: PlateResult, out_svg: Path, title: str = "") -> None:
    """16x24 grid of per-well traces with k_cat printed in each panel.

    Coloring per panel:
        * gray line   = raw A(t) trace
        * blue line   = baseline-corrected A(t) (raw minus water-control mean)
        * red shading = time window used for the initial-velocity fit
        * blue points = corrected data inside the fit window
        * blue line   = fitted initial-velocity slope (on corrected data)
    """
    plt.rcParams.update(NATURE_RC)
    fig, axes = plt.subplots(
        len(PLATE_ROWS), len(PLATE_COLS),
        figsize=(NATURE_PANEL_INCHES[0] * 1.5, NATURE_PANEL_INCHES[0] * 1.5 * (16 / 24)),
        sharex=True, sharey=True,
    )
    t = result.traces.times_s
    baseline = result.baseline_trace
    grid = result.kcat_grid() 
    vmax = np.nanpercentile(np.abs(grid), 95)
    if not np.isfinite(vmax) or vmax == 0:
        vmax = 1.0
    cmap = plt.get_cmap("viridis")
    fit_color = "#1f77b4"  # matplotlib default blue
    water_bg = "#dce9f7"   # light blue
    tcep_bg = "#d9ecd0"    # light green
    zprime_color = "#7a9cc6"
    zprime_end = min(float(t[-1]), Z_PRIME_WINDOW_S)

    for i, rlabel in enumerate(PLATE_ROWS):
        for j, clabel in enumerate(PLATE_COLS):
            ax = axes[i, j]
            w = f"{rlabel}{clabel}"
            a = result.traces.traces.get(w)
            fit = result.fits.get(w)
            excluded = w in result.excluded
            is_water = result.layout.is_water_control(w)
            is_tcep = result.layout.is_tcep_control(w)
            if excluded:
                ax.set_facecolor("#ffd6d6")
            elif is_water:
                ax.set_facecolor(water_bg)
            elif is_tcep:
                ax.set_facecolor(tcep_bg)
            if a is not None:
                ax.plot(t, a, color="0.55", lw=0.4)
                # Controls skip baseline correction; only draw corrected trace
                # for enzyme wells so the water/TCEP raw signal is what's shown.
                a_corr = a if (baseline is None or is_water or is_tcep) else a - baseline
                if baseline is not None and not (is_water or is_tcep):
                    ax.plot(t, a_corr, color=fit_color, lw=0.4, alpha=0.9)
                if fit is not None and np.isfinite(fit.v0_abs_per_s):
                    t0, t1 = fit.window_s
                    ax.axvspan(t0, t1, color="0.8", alpha=0.5, lw=0)
                    mask = (t >= t0) & (t <= t1)
                    ax.plot(t[mask], a_corr[mask], "o", ms=0.8,
                            color=fit_color, mec="none")
                    tl = np.array([t0, t1])
                    ax.plot(tl, fit.intercept + fit.v0_abs_per_s * tl,
                            color=fit_color, lw=0.7)
                    if np.isfinite(fit.kcat_s):
                        ax.text(
                            0.03, 0.03, f"{fit.kcat_s:.2g}",
                            transform=ax.transAxes, ha="left", va="bottom",
                            fontsize=4,
                            color=cmap(min(1.0, abs(fit.kcat_s) / vmax)),
                        )
                # Z' region drawn last on control wells so the gray fit-window
                # shading doesn't paint over it.
                if is_water or is_tcep:
                    ax.axvspan(float(t[0]), zprime_end,
                               color=zprime_color, alpha=0.35, lw=0)
            if excluded:
                ax.text(0.97, 0.97, "no enz",
                        transform=ax.transAxes, ha="right", va="top",
                        fontsize=4, color="crimson")
            ax.set_xticks([])
            ax.set_yticks([])
            for spine in ax.spines.values():
                spine.set_linewidth(0.3)
            if i == 0:
                ax.set_title(str(clabel), fontsize=4, pad=1)
            if j == 0:
                ax.set_ylabel(rlabel, fontsize=4, rotation=0, labelpad=4, va="center")

    # One shared legend at the top explains the coloring.
    from matplotlib.lines import Line2D
    from matplotlib.patches import Patch
    legend_handles = [
        Line2D([0], [0], color="0.55", lw=1.0, label="raw A(t)"),
        Line2D([0], [0], color=fit_color, lw=1.0, label="baseline-corrected A(t)"),
        Patch(facecolor="0.8", alpha=0.5, label="initial-velocity window"),
        Line2D([0], [0], marker="o", ms=3, color=fit_color, lw=1.0,
               label="fit points + v$_0$ slope"),
        Patch(facecolor=water_bg, label="water control"),
        Patch(facecolor=tcep_bg, label="TCEP control"),
        Patch(facecolor=zprime_color, alpha=0.35, label="Z' region"),
    ]
    fig.legend(handles=legend_handles, loc="upper center", ncol=7,
               frameon=False, fontsize=6, bbox_to_anchor=(0.5, 0.995))

    fig.suptitle(title or "Per-well traces; k$_{cat}$ (s$^{-1}$) printed in each well",
                 fontsize=7, y=0.965)
    z_raw = result.stats.get("z_prime_raw", float("nan"))
    if np.isfinite(z_raw):
        fig.text(0.995, 0.985, f"Z'$_{{raw}}$ = {z_raw:.2f}",
                 ha="right", va="top", fontsize=7)
    if result.rotation_applied:
        fig.text(0.005, 0.985, "plate rotated 180° (auto-corrected)",
                 ha="left", va="top", fontsize=7, color="crimson")
    fig.subplots_adjust(wspace=0.05, hspace=0.05, top=0.93, bottom=0.02,
                        left=0.03, right=0.99)
    out_svg.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out_svg, format="pdf")
    fig.savefig(out_svg.with_suffix(".png"), format="png", dpi=600)
    plt.close(fig)


def plot_kcat_heatmap(result: PlateResult, out_svg: Path, title: str = "") -> None:
    """Nature-panel-sized k_cat heat map with row/col condition labels."""
    plt.rcParams.update(NATURE_RC)
    fig, ax = plt.subplots(figsize=NATURE_PANEL_INCHES)
    grid = result.kcat_grid()

    finite = grid[np.isfinite(grid)]
    if finite.size:
        vmin = np.nanpercentile(finite, 2)
        vmax = np.nanpercentile(finite, 98)
    else:
        vmin, vmax = 0.0, 1.0
    cmap = plt.get_cmap("gray_r").copy()
    cmap.set_bad("#ff5252")
    im = ax.imshow(grid, cmap=cmap, aspect="auto", vmin=vmin, vmax=vmax)

    # Axis labels come from the layout (metals across, alcohols down).
    col_labels = [result.layout.metal.get(f"A{c}", "") for c in PLATE_COLS]
    row_labels = [result.layout.substrate.get(f"{r}1", "") for r in PLATE_ROWS]
    ax.set_xticks(range(len(PLATE_COLS)))
    ax.set_xticklabels(col_labels, rotation=90)
    ax.set_yticks(range(len(PLATE_ROWS)))
    ax.set_yticklabels(row_labels)
    ax.set_xlabel("metal (col)")
    ax.set_ylabel("substrate (row)")
    ax.set_title(title or "k$_{cat}$ (s$^{-1}$)")

    z_raw = result.stats.get("z_prime_raw", float("nan"))
    if np.isfinite(z_raw):
        ax.text(1.0, 1.02, f"Z'$_{{raw}}$ = {z_raw:.2f}",
                transform=ax.transAxes, ha="right", va="bottom", fontsize=7)
    if result.rotation_applied:
        ax.text(0.0, 1.02, "plate rotated 180° (auto-corrected)",
                transform=ax.transAxes, ha="left", va="bottom",
                fontsize=7, color="crimson")

    cb = fig.colorbar(im, ax=ax, shrink=0.85, pad=0.02)
    cb.ax.tick_params(labelsize=6)
    cb.outline.set_linewidth(0.4)
    cb.set_label("k$_{cat}$ (s$^{-1}$)")

    fig.tight_layout()
    out_svg.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out_svg, format="pdf")
    fig.savefig(out_svg.with_suffix(".png"), format="png", dpi=600)
    plt.close(fig)


# =============================================================================
# Multi-plate: averaging + cross-condition compare
# =============================================================================

def average_plate_results(results: list[PlateResult]) -> tuple[np.ndarray, np.ndarray]:
    """Return (mean, std) 16x24 k_cat grids across a list of replicate plates."""
    stack = np.stack([r.kcat_grid() for r in results], axis=0)
    return np.nanmean(stack, axis=0), np.nanstd(stack, axis=0)


def plot_averaged_heatmaps(
    mean_grid: np.ndarray,
    std_grid: np.ndarray,
    layout: PlateLayout,
    out_svg: Path,
    title: str = "",
) -> None:
    """Side-by-side mean / std k_cat heat maps for a triplicate average."""
    plt.rcParams.update(NATURE_RC)
    fig, axes = plt.subplots(1, 2, figsize=(NATURE_PANEL_INCHES[0] * 2, NATURE_PANEL_INCHES[1]))
    col_labels = [layout.metal.get(f"A{c}", "") for c in PLATE_COLS]
    row_labels = [layout.substrate.get(f"{r}1", "") for r in PLATE_ROWS]

    for ax, grid, label in zip(axes, [mean_grid, std_grid], ["mean k$_{cat}$", "std k$_{cat}$"]):
        finite = grid[np.isfinite(grid)]
        vmin = np.nanpercentile(finite, 2) if finite.size else 0
        vmax = np.nanpercentile(finite, 98) if finite.size else 1
        cmap = plt.get_cmap("gray_r").copy()
        cmap.set_bad("#ff5252")
        im = ax.imshow(grid, cmap=cmap, vmin=vmin, vmax=vmax, aspect="auto")
        ax.set_xticks(range(len(PLATE_COLS))); ax.set_xticklabels(col_labels, rotation=90)
        ax.set_yticks(range(len(PLATE_ROWS))); ax.set_yticklabels(row_labels)
        ax.set_title(label)
        cb = fig.colorbar(im, ax=ax, shrink=0.85, pad=0.02)
        cb.ax.tick_params(labelsize=6); cb.outline.set_linewidth(0.4)
    if title:
        fig.suptitle(title, fontsize=7)
    fig.tight_layout()
    out_svg.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out_svg, format="pdf")
    fig.savefig(out_svg.with_suffix(".png"), format="png", dpi=600)
    plt.close(fig)


def compare_conditions(
    condition_grids: dict[str, np.ndarray],
    layout: PlateLayout,
    out_svg: Path,
    out_csv: Path,
) -> None:
    """Render a grid of k_cat heat maps, one per (protein, pH) condition, and dump
    a tidy CSV: condition,well,row_label,col_label,kcat.
    """
    plt.rcParams.update(NATURE_RC)
    n = len(condition_grids)
    ncols = min(n, 3)
    nrows = int(np.ceil(n / ncols))
    fig, axes = plt.subplots(nrows, ncols,
                             figsize=(NATURE_PANEL_INCHES[0] * ncols,
                                      NATURE_PANEL_INCHES[1] * nrows),
                             squeeze=False)

    all_vals = np.concatenate([g[np.isfinite(g)].ravel() for g in condition_grids.values()])
    if all_vals.size:
        vmin = float(np.nanpercentile(all_vals, 2))
        vmax = float(np.nanpercentile(all_vals, 98))
    else:
        vmin, vmax = 0.0, 1.0

    col_labels = [layout.metal.get(f"A{c}", "") for c in PLATE_COLS]
    row_labels = [layout.substrate.get(f"{r}1", "") for r in PLATE_ROWS]
    for k, (name, grid) in enumerate(condition_grids.items()):
        ax = axes[k // ncols][k % ncols]
        cmap = plt.get_cmap("gray_r").copy()
        cmap.set_bad("#ff5252")
        im = ax.imshow(grid, cmap=cmap, vmin=vmin, vmax=vmax, aspect="auto")
        ax.set_xticks(range(len(PLATE_COLS))); ax.set_xticklabels(col_labels, rotation=90)
        ax.set_yticks(range(len(PLATE_ROWS))); ax.set_yticklabels(row_labels)
        ax.set_title(name)
        cb = fig.colorbar(im, ax=ax, shrink=0.85, pad=0.02)
        cb.ax.tick_params(labelsize=6); cb.outline.set_linewidth(0.4)

    for extra in range(len(condition_grids), nrows * ncols):
        axes[extra // ncols][extra % ncols].axis("off")

    fig.suptitle(f"k$_{{cat}}$ across {n} conditions (shared colour scale)", fontsize=7)
    fig.tight_layout()
    out_svg.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out_svg, format="pdf")
    fig.savefig(out_svg.with_suffix(".png"), format="png", dpi=600)
    plt.close(fig)

    with out_csv.open("w", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(["condition", "well", "row_label", "col_label", "kcat_s"])
        for name, grid in condition_grids.items():
            for i, r in enumerate(PLATE_ROWS):
                for j, c in enumerate(PLATE_COLS):
                    well = f"{r}{c}"
                    w.writerow([name, well,
                                layout.substrate.get(well, ""),
                                layout.metal.get(well, ""),
                                grid[i, j]])


# =============================================================================
# Single-plate driver + CSV/stats dump
# =============================================================================

# Wells whose ΔA (start-plateau) is not clearly above the water-control distribution
# are treated as "no enzyme added" and excluded from downstream stats/heatmap.
AUTO_EXCLUDE_SIGMA: float = 3.0

# Seconds from t = 0 used as the raw-Abs region for the plate-quality Z'.
Z_PRIME_WINDOW_S: float = 300.0


def _rotate_well_id_180(well: str) -> str:
    """Return the well id at the 180-degree partner position (A1 <-> P24 etc.)."""
    row_idx = PLATE_ROWS.index(well[0].upper())
    col = int(well[1:])
    new_row = PLATE_ROWS[len(PLATE_ROWS) - 1 - row_idx]
    new_col = len(PLATE_COLS) + 1 - col
    return f"{new_row}{new_col}"


def _rotate_traces_180(traces: PlateTraces) -> PlateTraces:
    remapped = {_rotate_well_id_180(w): v for w, v in traces.traces.items()}
    return PlateTraces(times_s=traces.times_s, traces=remapped,
                       temperature_c=traces.temperature_c, source=traces.source)


def _row_lr_split(traces: PlateTraces, row: str, mask: np.ndarray) -> float:
    """mean(cols 1-12) - mean(cols 13-24) of the first-window absorbance for `row`."""
    left, right = [], []
    for c in PLATE_COLS:
        a = traces.traces.get(f"{row}{c}")
        if a is None:
            continue
        sub = a[mask]
        if not np.any(np.isfinite(sub)):
            continue
        val = float(np.nanmean(sub))
        (left if c <= 12 else right).append(val)
    if not left or not right:
        return float("nan")
    return float(np.mean(left) - np.mean(right))


def _looks_rotated_180(traces: PlateTraces) -> bool:
    """Heuristic: on the test-plate layout, row P should carry the water|TCEP split
    across cols 1-12 vs 13-24. If row A shows a big left/right split and row P does
    not, the plate was probably loaded 180-degrees rotated."""
    mask = traces.times_s <= Z_PRIME_WINDOW_S
    if not np.any(mask):
        mask = np.ones_like(traces.times_s, dtype=bool)
    p_split = _row_lr_split(traces, "P", mask)
    a_split = _row_lr_split(traces, "A", mask)
    if not (np.isfinite(p_split) and np.isfinite(a_split)):
        return False
    return abs(a_split) > 3.0 * max(abs(p_split), 0.05)


def _auto_detect_bad_wells(
    traces: PlateTraces,
    layout: PlateLayout,
    *,
    sigma: float = AUTO_EXCLUDE_SIGMA,
    edge_frac: float = 0.1,
) -> set[str]:
    """Flag enzyme wells whose total absorbance drop is indistinguishable
    from the water (negative) control distribution.

    ΔA is the mean of the first ``edge_frac`` of the trace minus the mean of
    the last ``edge_frac`` (positive when DCPIP is being reduced). A well is
    flagged when its ΔA is below ``water_mean + sigma * water_std``.
    """
    def _delta_a(w: str) -> float:
        a = traces.traces.get(w)
        if a is None or len(a) < 4:
            return float("nan")
        n = max(3, int(round(len(a) * edge_frac)))
        head, tail = a[:n], a[-n:]
        if not np.any(np.isfinite(head)) or not np.any(np.isfinite(tail)):
            return float("nan")
        return float(np.nanmean(head) - np.nanmean(tail))

    water = np.array([_delta_a(w) for w in traces.traces
                      if layout.is_water_control(w)])
    water = water[np.isfinite(water)]
    if water.size < 3:
        return set()
    threshold = float(np.mean(water) + sigma * np.std(water))

    bad: set[str] = set()
    for w in traces.traces:
        if (layout.is_water_control(w) or layout.is_tcep_control(w)
                or layout.is_no_substrate(w) or layout.is_no_metal(w)):
            continue
        d = _delta_a(w)
        if not np.isfinite(d) or d < threshold:
            bad.add(w)
    return bad


def analyze_plate(
    xlsx_path: Path,
    *,
    eps_app: float,
    enzyme_uM: float,
    layout: PlateLayout,
    window_s: float = 300.0,
    rotate: str = "auto",
) -> PlateResult:
    traces = parse_kinetic_plate(xlsx_path)
    detected = _looks_rotated_180(traces)
    if rotate == "180":
        apply_rotation = True
    elif rotate == "off":
        apply_rotation = False
    else:  # "auto"
        apply_rotation = detected
    if apply_rotation:
        traces = _rotate_traces_180(traces)
    fits = compute_well_fits(traces, layout, eps_app, enzyme_uM, window_s=window_s)
    excluded = _auto_detect_bad_wells(traces, layout)
    for w in excluded:
        f = fits.get(w)
        if f is not None:
            f.kcat_s = float("nan")
    result = PlateResult(traces=traces, layout=layout, fits=fits,
                         eps_app=eps_app, enzyme_uM=enzyme_uM, excluded=excluded,
                         baseline_trace=_water_baseline_trace(traces, layout),
                         rotate_mode=rotate,
                         rotation_detected=detected,
                         rotation_applied=apply_rotation)
    result.stats = compute_plate_stats(result)
    return result


def dump_well_table(result: PlateResult, out_csv: Path) -> None:
    out_csv.parent.mkdir(parents=True, exist_ok=True)
    with out_csv.open("w", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(["well", "row", "col", "metal", "substrate",
                    "v0_abs_per_s", "kcat_s", "r_squared",
                    "window_start_s", "window_end_s", "n_points"])
        for r in PLATE_ROWS:
            for c in PLATE_COLS:
                well = f"{r}{c}"
                fit = result.fits.get(well)
                metal = result.layout.metal.get(well, "")
                sub = result.layout.substrate.get(well, "")
                if fit is None:
                    w.writerow([well, r, c, metal, sub, "", "", "", "", "", ""])
                    continue
                w.writerow([well, r, c, metal, sub,
                            fit.v0_abs_per_s, fit.kcat_s, fit.r_squared,
                            fit.window_s[0], fit.window_s[1], fit.n_points])


def dump_stats(result: PlateResult, out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w") as fh:
        fh.write(f"# Plate stats for {result.traces.source}\n")
        fh.write(f"eps_app (AU/µM): {result.eps_app}\n")
        fh.write(f"[E] (µM):        {result.enzyme_uM}\n")
        fh.write(f"layout:          {result.layout.description}\n")
        fh.write(f"rotate mode:     {result.rotate_mode}\n")
        fh.write(f"rotation:        detected={result.rotation_detected}, "
                 f"applied={result.rotation_applied}\n")
        if result.excluded:
            fh.write(f"auto-excluded:   {', '.join(sorted(result.excluded))}\n")
        fh.write("\n")
        for k, v in result.stats.items():
            fh.write(f"{k:>22s}: {v:.6g}\n")


# =============================================================================
# eps_app loader
# =============================================================================

def load_eps_app(explicit: float | None, standard_curve_xlsx: Path | None) -> float:
    if explicit is not None:
        return float(explicit)
    if standard_curve_xlsx is not None:
        from dcpip_standard_curve import fit_standard_curve  # local import to avoid cycles
        fit = fit_standard_curve(standard_curve_xlsx)
        print(f"[standard curve] eps_app = {fit.eps_app:.6g} AU/µM  R^2 = {fit.r_squared:.4f}")
        return fit.eps_app
    return DEFAULT_EPS_APP_AU_PER_UM


if __name__ == "__main__":
    main()
